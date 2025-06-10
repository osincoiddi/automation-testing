import openpyxl
from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter


def getRowCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.max_column

def fillCell(file,sheetName, row, col, value):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=row,column=col,value=value)
    wb.save(file)
    return "Done"

def getReadData(file,sheetName,row,col):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.cell(row=row,column=col).value
def getWriteData(file,sheetName,row,col):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.cell(row=row,column=col).value
def fillCellColor(file,sheetName,row,col,color):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=row,column=col).fill=PatternFill(start_color=color,end_color=color,fill_type='solid')
    wb.save(file)
    return "Done"
def getFillGreenColor(file,sheetName,row,col):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    greenfile=PatternFill(start_color='60b1212',end_color='60b1212',fill_type='solid')
    sheet.cell(row=row,column=col).fill=greenfile
    wb.save(file)
    return "Done"

def getFillRedColor(file,sheetName,row,col):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    redfile=PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    sheet.cell(row=row,column=col).fill=redfile
    wb.save(file)
    return "Done"






